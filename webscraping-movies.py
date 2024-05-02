
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2022/'

##   Open the website in python
page = urlopen(webpage)			
soup = BeautifulSoup(page, 'html.parser')
title = soup.title
print(title.text)
table_rows = soup.findAll("tr")

##set up the excel workbook
wb = xl.Workbook()
ws = wb.active
ws.title = 'Box Office Report'
write_sheet = wb['Box Office Report']

## Create Column Titles
ws['A1'] = 'No.'
ws['A1'].font = Font(name='Times New Roman', size=20, italic=False, bold=True)

ws['B1'] = 'Movie Title'
ws['B1'].font = Font(name='Times New Roman', size=20, italic=False, bold=True)

ws['C1'] = 'Release Date'
ws['C1'].font = Font(name='Times New Roman', size=20, italic=False, bold=True)

ws['D1'] = 'Total Gross'
ws['D1'].font = Font(name='Times New Roman', size=20, italic=False, bold=True)

ws['E1'] = 'Theaters'
ws['E1'].font = Font(name='Times New Roman', size=20, italic=False, bold=True)

ws['F1'] = 'Average per Theater'
ws['F1'].font = Font(name='Times New Roman', size=20, italic=False, bold=True)

##go through the rows and add information to the sheet
cell = 2

for row in table_rows[1:6]:
    ## NEEDED this will find the next column in the table
    td = soup.findAll("td")

    ## pull the data from website into variables
    rank = int(td[0].text)
    movie = td[1].text
    release_date = td[8].text
    gross_rev = float(td[7].text.replace(",", "").replace("$", ""))
    theaters = int(td[6].text.replace(",", ""))

    ## put variables into the excel sheet
    write_sheet.cell(cell,1).value = rank
    write_sheet.cell(cell,2).value = movie
    write_sheet.cell(cell,3).value = release_date
    write_sheet.cell(cell,4).value = theaters
    write_sheet.cell(cell,5).value = gross_rev
    write_sheet.cell(cell,6).value = '=E' + str(cell) + '/D' + str(cell)

    cell += 1


## adjusts the width of the columns in the table
write_sheet.column_dimensions['A'].width = 5
write_sheet.column_dimensions['B'].width = 45
write_sheet.column_dimensions['C'].width = 30
write_sheet.column_dimensions['D'].width = 20
write_sheet.column_dimensions['E'].width = 20
write_sheet.column_dimensions['F'].width = 45

## format
for cell in write_sheet["D:D"]:
    cell.number_format = u'"$ "#,##0.00'

for cell in write_sheet["E:E"]:
    cell.number_format = '#,##0'

for cell in write_sheet["F:F"]:
    cell.number_format = u'"$ "#,##0.00'


## saves the workbook. 
## This will cause an error if you have the file open
wb.save('BoxOfficeReport.xlsx')


##
##
##

